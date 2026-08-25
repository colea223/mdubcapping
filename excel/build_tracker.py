"""
Builds the MW Handicapping Tracker.xlsx workbook: Read Me, Settings,
Weekly Slate, Bet Log, and Team Profiles tabs.

Run with the project's normal Python (openpyxl/pandas already present):
    python3 excel/build_tracker.py
Then recalc (mandatory -- openpyxl writes formulas with no cached values):
    python3 /mnt/skills/public/docx/scripts/office/soffice.py --headless --convert-to xlsx ... # n/a
    python3 <xlsx skill dir>/scripts/recalc.py "MW Handicapping Tracker.xlsx"
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from teams import MW_TEAMS_2026  # noqa: E402

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=10)          # blue = user input
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)        # black = formula
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")             # yellow = key assumption
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="666666")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------------------------------------------------------- Read Me
ws = wb.active
ws.title = "Read Me"
ws["A1"] = "Mountain West Handicapping Tracker"
ws["A1"].font = TITLE_FONT
ws["A2"] = "2026 Season"
ws["A2"].font = NOTE_FONT

lines = [
    "",
    "How this workbook is meant to be used",
    "This workbook is the review/tracking cockpit -- the model itself runs in Python (see the",
    "attack plan and the src/ scripts). Each week, paste that week's model output into the blue",
    "cells on the Weekly Slate tab; everything else on that tab recalculates automatically.",
    "",
    "Color legend",
    "  Blue text      = cells you fill in by hand (model output, market lines, bet results)",
    "  Black text      = formulas -- don't type over these, they recalculate from the blue cells",
    "  Yellow fill     = key assumptions on the Settings tab (edit these once, not weekly)",
    "",
    "Tabs",
    "  Settings        = bankroll, edge thresholds, and other assumptions used across the workbook",
    "  Weekly Slate    = this week's matchups: model line vs. market line, edge, and a suggested lean",
    "  Bet Log         = every bet placed -- stake, line, closing line (for CLV), result, running bankroll",
    "  Team Profiles   = one row per 2026 Mountain West team, pre-filled with conference history;",
    "                    paste in power ratings / PPA / talent rank weekly as your pipeline produces them",
    "",
    "This is a tracking tool, not a guarantee of profit -- see the attack plan's closing note on",
    "paper-trading and bankroll discipline before staking real money.",
]
for i, line in enumerate(lines, start=3):
    cell = ws.cell(row=i, column=1, value=line)
    cell.font = Font(name=FONT_NAME, size=10, bold=line.strip() in (
        "How this workbook is meant to be used", "Color legend", "Tabs"
    ))
autosize(ws, [110])

# ---------------------------------------------------------------- Settings
ws = wb.create_sheet("Settings")
ws["A1"] = "Settings & Assumptions"
ws["A1"].font = TITLE_FONT
settings_rows = [
    ("Starting bankroll (units)", 100, "Example: 100 units. 1 unit = whatever flat stake you choose in real currency."),
    ("Spread edge threshold (pts)", 2.0, "Minimum |model line - market line| before a spread is flagged as a lean."),
    ("Total edge threshold (pts)", 2.0, "Minimum |model total - market total| before a total is flagged as a lean."),
    ("Standard vig odds (American)", -110, "Reference only -- breakeven win rate at -110 is 52.4%."),
]
ws["A3"] = "Assumption"
ws["B3"] = "Value"
ws["C3"] = "Notes"
style_header_row(ws, 3, 3)
for i, (label, value, note) in enumerate(settings_rows, start=4):
    ws.cell(row=i, column=1, value=label).font = FORMULA_FONT
    vcell = ws.cell(row=i, column=2, value=value)
    vcell.font = INPUT_FONT
    vcell.fill = ASSUMPTION_FILL
    ws.cell(row=i, column=3, value=note).font = NOTE_FONT
autosize(ws, [32, 12, 70])
# Named references for use in formulas on other sheets
BANKROLL_CELL = "Settings!$B$4"
SPREAD_THRESH_CELL = "Settings!$B$5"
TOTAL_THRESH_CELL = "Settings!$B$6"

# ---------------------------------------------------------------- Weekly Slate
ws = wb.create_sheet("Weekly Slate")
headers = [
    "Week", "Date", "Away Team", "Home Team",
    "Model Line (Home)", "Market Line (Home)", "Spread Edge (pts)", "Spread Lean",
    "Model Total", "Market Total", "Total Edge (pts)", "Total Lean",
    "Suggested Bet (Spread)", "Suggested Bet (Total)", "Qualitative Notes", "Final Decision",
]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))

# One example row (row 2) with realistic values, per the xlsx skill's requirement
# that a fill-in workbook show one worked example.
example = {
    "A": 1, "B": "2026-08-29", "C": "San José State", "D": "Air Force",
    "E": -6.5, "F": -4.5,
}
for col, val in example.items():
    cell = ws[f"{col}2"]
    cell.value = val
    cell.font = INPUT_FONT

ws["G2"] = "=F2-E2"                                            # Spread Edge = Market - Model
ws["H2"] = '=IF(G2=0,"Pick\'em",IF(G2>0,"Home","Away"))'
ws["I2"] = 44.5
ws["I2"].font = INPUT_FONT
ws["J2"] = 41.0
ws["J2"].font = INPUT_FONT
ws["K2"] = "=I2-J2"                                             # Total Edge = Model - Market
ws["L2"] = '=IF(K2=0,"Push",IF(K2>0,"Over","Under"))'
ws["M2"] = f'=IF(ABS(G2)>={SPREAD_THRESH_CELL},H2,"Pass")'
ws["N2"] = f'=IF(ABS(K2)>={TOTAL_THRESH_CELL},L2,"Pass")'
ws["O2"] = "Starting QB confirmed healthy per Thu injury report"
ws["O2"].font = INPUT_FONT
ws["P2"] = "Home -4.5"
ws["P2"].font = INPUT_FONT
for col in ("G", "H", "K", "L", "M", "N"):
    ws[f"{col}2"].font = FORMULA_FONT

# 40 more blank rows ready to fill in, formulas pre-filled down the column
for r in range(3, 43):
    ws[f"G{r}"] = f"=F{r}-E{r}"
    ws[f"H{r}"] = f'=IF(G{r}=0,"Pick\'em",IF(G{r}>0,"Home","Away"))'
    ws[f"K{r}"] = f"=I{r}-J{r}"
    ws[f"L{r}"] = f'=IF(K{r}=0,"Push",IF(K{r}>0,"Over","Under"))'
    ws[f"M{r}"] = f'=IF(ABS(G{r})>={SPREAD_THRESH_CELL},H{r},"Pass")'
    ws[f"N{r}"] = f'=IF(ABS(K{r})>={TOTAL_THRESH_CELL},L{r},"Pass")'
    for col in ("G", "H", "K", "L", "M", "N"):
        ws[f"{col}{r}"].font = FORMULA_FONT

autosize(ws, [6, 11, 15, 15, 16, 16, 14, 11, 11, 11, 12, 10, 17, 16, 30, 16])
ws.freeze_panes = "A2"

# ---------------------------------------------------------------- Bet Log
ws = wb.create_sheet("Bet Log")
headers = [
    "Date", "Week", "Matchup", "Bet Type", "Side", "Line Taken", "Odds (American)",
    "Stake (units)", "Closing Line", "CLV (pts)", "Result", "Units Won/Lost", "Running Bankroll",
]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))

example_bet = {
    "A": "2026-08-29", "B": 1, "C": "San José State @ Air Force", "D": "Spread",
    "E": "Air Force -4.5", "F": -4.5, "G": -110, "H": 2, "I": -6.0,
}
for col, val in example_bet.items():
    cell = ws[f"{col}2"]
    cell.value = val
    cell.font = INPUT_FONT
ws["J2"] = "=I2-F2"   # CLV: closing line minus line taken (see Read Me note on sign convention)
ws["K2"] = "W"
ws["K2"].font = INPUT_FONT
ws["L2"] = '=IF(K2="W",IF(G2<0,H2*100/ABS(G2),H2*G2/100),IF(K2="L",-H2,0))'
ws["M2"] = f"={BANKROLL_CELL}+SUM(L$2:L2)"
for col in ("J", "L", "M"):
    ws[f"{col}2"].font = FORMULA_FONT

for r in range(3, 43):
    ws[f"J{r}"] = f"=IF(I{r}=\"\",\"\",I{r}-F{r})"
    ws[f"L{r}"] = (
        f'=IF(K{r}="W",IF(G{r}<0,H{r}*100/ABS(G{r}),H{r}*G{r}/100),'
        f'IF(K{r}="L",-H{r},0))'
    )
    ws[f"M{r}"] = f"={BANKROLL_CELL}+SUM(L$2:L{r})"
    for col in ("J", "L", "M"):
        ws[f"{col}{r}"].font = FORMULA_FONT

autosize(ws, [11, 6, 26, 10, 18, 11, 14, 12, 12, 10, 9, 14, 15])
ws.freeze_panes = "A2"
ws["A45"] = "CLV note: Closing Line - Line Taken. For a favorite (negative number), a less-negative closing line than what you bet means the market moved toward you -- positive CLV. Read each row in context of which side you took."
ws["A45"].font = NOTE_FONT

# ---------------------------------------------------------------- Team Profiles
ws = wb.create_sheet("Team Profiles")
headers = [
    "Team", "Prior Conference", "Joined 2026", "Model Power Rating", "SP+ Rating",
    "Off PPA", "Def PPA", "Talent Rank", "Notes",
]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))

for i, (team, meta) in enumerate(sorted(MW_TEAMS_2026.items()), start=2):
    ws.cell(row=i, column=1, value=team).font = FORMULA_FONT
    ws.cell(row=i, column=2, value=meta["prior_conference"]).font = FORMULA_FONT
    ws.cell(row=i, column=3, value="Y" if meta["joined_2026"] else "N").font = FORMULA_FONT
    for col in (4, 5, 6, 7, 8):
        c = ws.cell(row=i, column=col)
        c.font = INPUT_FONT  # blank input cells -- paste weekly pipeline output here
    ws.cell(row=i, column=9, value=meta["notes"]).font = NOTE_FONT

autosize(ws, [16, 26, 11, 17, 12, 10, 10, 11, 60])
ws.freeze_panes = "A2"

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

out_path = Path(__file__).resolve().parent / "MW_Handicapping_Tracker.xlsx"
wb.save(out_path)
print(f"Saved {out_path}")
