"""
Adds/refreshes a "Model Comparison" tab in MW_Handicapping_Tracker.xlsx --
Ridge (the live model) vs. XGBoost (a candidate), graded against Vegas's
closing spread, side by side. Reads src/model_comparison.py's output
(data/clean/model_comparison_results.csv -- run that script first) for the
historical/graded sections, PLUS src/predict_week.py's latest predictions
CSV for a live "Upcoming Games" section (Ridge's and XGBoost's current
lines for the next not-yet-played week, ungraded since there's no result
yet) -- see write_upcoming_block(). That section also needs a database
connection (for the current market line on each upcoming game), which is
why main() now opens one against db/mw_handicapping.duckdb.

NON-DESTRUCTIVE, same rule as excel/update_tracker.py: this only ever
touches the "Model Comparison" sheet -- creating it if it doesn't exist yet,
or clearing and rewriting ONLY that sheet's cells if it does. Every other
tab (Read Me, Settings, Weekly Slate, Bet Log, Team Profiles, and anything
else you've added) is left completely untouched.

This does NOT change what Weekly Slate uses, and does NOT touch Ridge's
role as the live model anywhere in this project -- it's an informational
side-by-side, nothing more, per the explicit instruction that XGBoost isn't
taking over yet.

CLV (closing-line value) columns -- a separate axis from ATS win/loss above:
whichever model's own predicted line ends up numerically CLOSER to the
actual closing market number is "tighter" that game (see _clv_stats()'s and
write_detail_table()'s docstrings), and each model's signed edge accumulates
into a running "CLV Pts" total, in raw points (not converted to units/
dollars). This is purely a numerical-accuracy contest against the closing
line -- independent of which side either model favored or whether a bet was
made. The Upcoming Games table gets a provisional "Tighter (Live)*" version
of the same idea, against today's still-moving market line, clearly marked
as not-final.

Usage:
    source .venv/bin/activate     (or the Windows equivalent)
    python src/predict_week.py              # produces the upcoming-games CSV this reads
    python src/model_comparison.py          # produces the graded-history CSV this reads
    python excel/update_model_comparison_tab.py
"""
import re
import sys
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from config import CLEAN_DIR, DB_PATH  # noqa: E402

TRACKER_PATH = Path(__file__).resolve().parent / "MW_Handicapping_Tracker.xlsx"
RESULTS_PATH = CLEAN_DIR / "model_comparison_results.csv"
SHEET_NAME = "Model Comparison"

PRED_FILE_RE = re.compile(r"^week_(\d{4})_(\d+)_predictions\.csv$")

# Same visual language as build_tracker.py's other tabs, redefined here
# rather than imported -- build_tracker.py runs (and saves a fresh workbook)
# at import time, so importing it would clobber the real tracker.
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F3864")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="666666")
WIN_FILL = PatternFill("solid", fgColor="E2EFDA")
LOSS_FILL = PatternFill("solid", fgColor="FCE4EC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _clv_stats(df_slice, prefix):
    """
    CLV (closing-line value) for one model over a slice of graded games --
    NOT the same thing as ATS win/loss above. This compares the model's own
    predicted line straight to the actual closing market line, regardless of
    which side either model favored or whether a bet was even made:
      - "tighter" (a game-by-game contest): whichever model's predicted line
        ends up numerically CLOSER to where the market actually closed --
        i.e. smaller |model_line - closing_line|. Since model_comparison.py's
        own "edge" column is already defined as (closing_line - model_line),
        |edge| IS that same distance -- no new quantity, just compared
        differently than edge is used elsewhere (picking a bet side).
      - "CLV points" (a running total): the SIGNED edge, summed across every
        graded game in this slice, in raw points -- not converted to units/
        dollars. Confirmed with Cole this is meant as a raw point tally, same
        units the per-game Edge columns already use, not a real-money figure.
    Requires the OTHER model's edge too (to know who was tighter each game),
    so this always looks at both ridge_edge and xgb_edge together.
    """
    edge_col = f"{prefix}_edge"
    other_prefix = "xgb" if prefix == "ridge" else "ridge"
    other_edge_col = f"{other_prefix}_edge"
    valid = df_slice.dropna(subset=[edge_col, other_edge_col])
    if valid.empty:
        return {"tighter_games": 0, "tighter_pct": None, "clv_pts": 0.0, "n": 0}
    tighter_games = int((valid[edge_col].abs() < valid[other_edge_col].abs()).sum())
    return {
        "tighter_games": tighter_games,
        "tighter_pct": round(tighter_games / len(valid), 4),
        "clv_pts": round(float(valid[edge_col].sum()), 1),
        "n": len(valid),
    }


def _summary_row(ws, row, name, s, clv):
    n_bets = s.get("n_bets") or 0
    ws.cell(row=row, column=1, value=name).font = FORMULA_FONT
    ws.cell(row=row, column=2, value=s.get("n_games", 0)).font = FORMULA_FONT
    ws.cell(row=row, column=3, value=n_bets).font = FORMULA_FONT
    ws.cell(row=row, column=4, value=s.get("wins")).font = FORMULA_FONT
    ws.cell(row=row, column=5, value=s.get("losses")).font = FORMULA_FONT
    ws.cell(row=row, column=6, value=s.get("pushes")).font = FORMULA_FONT
    win_pct = s.get("ats_win_rate")
    ws.cell(row=row, column=7, value=(f"{win_pct:.1%}" if win_pct is not None else "n/a")).font = FORMULA_FONT
    roi = s.get("roi_flat_stake")
    ws.cell(row=row, column=8, value=(f"{roi:+.1%}" if roi is not None else "n/a")).font = FORMULA_FONT
    tighter_pct = clv.get("tighter_pct")
    ws.cell(row=row, column=9,
            value=(f"{clv.get('tighter_games', 0)}/{clv.get('n', 0)} ({tighter_pct:.0%})"
                   if tighter_pct is not None else "n/a")).font = FORMULA_FONT
    ws.cell(row=row, column=10, value=f"{clv.get('clv_pts', 0.0):+.1f}").font = FORMULA_FONT


def write_summary_block(ws, df, start_row, title, season_filter=None):
    """One 'Overall / MW-involved x Ridge / XGBoost' block. Returns the next free row."""
    from model_comparison import summarize

    sl = df if season_filter is None else df[df["season"] == season_filter]
    r = start_row
    ws.cell(row=r, column=1, value=title).font = SECTION_FONT
    r += 1
    headers = ["Model", "Games Graded", "Bets Made", "Wins", "Losses", "Pushes", "ATS Win %",
               "ROI (flat stake)", "Tighter Than Other (CLV)", "CLV Pts (cum.)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    r += 1

    for prefix, label in [("ridge", "Ridge (live model)"), ("xgb", "XGBoost (candidate)")]:
        s_overall = summarize(sl, prefix, "overall")
        clv_overall = _clv_stats(sl, prefix)
        _summary_row(ws, r, f"{label} -- all FBS", s_overall, clv_overall)
        r += 1
        mw_sl = sl[sl["is_mw_game"]]
        s_mw = summarize(mw_sl, prefix, "mw")
        clv_mw = _clv_stats(mw_sl, prefix)
        _summary_row(ws, r, f"{label} -- MW-involved", s_mw, clv_mw)
        r += 1
    return r + 1


def write_detail_table(ws, df, start_row, season):
    """Per-game detail for the CURRENT SEASON's Mountain West games only --
    same scoping convention as Weekly Slate/Bet Log (this season, MW focus),
    keeping the sheet a quick weekly read rather than an 11-season dump. The
    full history stays in data/clean/model_comparison_results.csv.

    Also carries the CLV columns from this same slice: "Tighter CLV Line"
    (which model's predicted line ended up closer to the actual closing
    market number this game -- see _clv_stats()'s docstring) and a running
    per-model "CLV Pts (Cum.)" total, accumulated top-to-bottom in this
    table's own row order (week ascending) -- same running-total convention
    as the Bet Log tab's Running Bankroll column, just for CLV points
    instead of bet units.
    """
    mw_this_season = df[(df["season"] == season) & (df["is_mw_game"])].sort_values(
        ["week", "home_team"]
    )

    r = start_row
    ws.cell(row=r, column=1, value=f"{season} Mountain West Games -- Game by Game").font = SECTION_FONT
    r += 1
    headers = [
        "Week", "Matchup", "Final Score", "Vegas Spread (Home)", "Actual Margin (Home)",
        "Ridge Line", "Ridge Edge", "Ridge Lean", "Ridge Result",
        "XGBoost Line", "XGBoost Edge", "XGBoost Lean", "XGBoost Result",
        "Models Agree?", "Tighter CLV Line", "Ridge CLV Pts (Cum.)", "XGBoost CLV Pts (Cum.)",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    header_row = r
    r += 1

    if mw_this_season.empty:
        ws.cell(row=r, column=1, value="No graded Mountain West games yet this season.").font = NOTE_FONT
        return r + 1, header_row

    def fmt_spread(v):
        if v is None or pd.isna(v):
            return "--"
        return f"+{v:.1f}" if v > 0 else f"{v:.1f}"

    def result_text(result, lean):
        # A CSV round-trip turns a missing (never-bet) result into NaN, not
        # Python None -- and `nan or fallback` is a trap (NaN is truthy in
        # Python), so this checks pd.isna() explicitly rather than relying
        # on plain truthiness.
        if isinstance(result, str) and result and not pd.isna(result):
            return result
        return "Lean only" if lean != "Pick'em" else "--"

    def fmt_score(row):
        # Same Away @ Home order as the Matchup column, so the two read
        # together naturally (e.g. "Wyoming @ Boise State" / "20-34").
        if pd.isna(row.away_points) or pd.isna(row.home_points):
            return "--"
        return f"{int(row.away_points)}-{int(row.home_points)}"

    ridge_clv_cum = 0.0
    xgb_clv_cum = 0.0
    for row in mw_this_season.itertuples():
        matchup = f"{row.away_team} @ {row.home_team}"
        ridge_result_text = result_text(row.ridge_result, row.ridge_lean)
        xgb_result_text = result_text(row.xgb_result, row.xgb_lean)

        # Tighter CLV Line: whichever model's predicted line ended up
        # numerically closer to the actual closing market number, i.e.
        # smaller |edge| (edge is already market_close - model_line) --
        # unrelated to which side either model favored or whether it was a
        # bet. Running totals accumulate the SIGNED edge, in this table's own
        # row order (week ascending), same convention as Bet Log's Running
        # Bankroll column -- see this function's docstring.
        if pd.notna(row.ridge_edge) and pd.notna(row.xgb_edge):
            ridge_abs, xgb_abs = abs(row.ridge_edge), abs(row.xgb_edge)
            tighter = "Ridge" if ridge_abs < xgb_abs else ("XGBoost" if xgb_abs < ridge_abs else "Tie")
        else:
            tighter = "n/a"
        if pd.notna(row.ridge_edge):
            ridge_clv_cum += row.ridge_edge
        if pd.notna(row.xgb_edge):
            xgb_clv_cum += row.xgb_edge

        values = [
            int(row.week), matchup, fmt_score(row),
            fmt_spread(row.market_spread_home), fmt_spread(row.actual_margin),
            fmt_spread(row.ridge_spread_home), fmt_spread(row.ridge_edge),
            row.ridge_lean, ridge_result_text,
            fmt_spread(row.xgb_spread_home), fmt_spread(row.xgb_edge),
            row.xgb_lean, xgb_result_text,
            ("Yes" if row.models_agree is True else ("No" if row.models_agree is False else "n/a")),
            tighter, f"{ridge_clv_cum:+.1f}", f"{xgb_clv_cum:+.1f}",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = FORMULA_FONT
            if c == 9 and row.ridge_result in ("Win", "Loss"):
                cell.fill = WIN_FILL if row.ridge_result == "Win" else LOSS_FILL
            if c == 13 and row.xgb_result in ("Win", "Loss"):
                cell.fill = WIN_FILL if row.xgb_result == "Win" else LOSS_FILL
            if c == 15 and tighter in ("Ridge", "XGBoost"):
                cell.fill = WIN_FILL
        r += 1

    return r + 1, header_row


def latest_predictions_file():
    """Same convention as excel/update_tracker.py's own helper of this name --
    duplicated rather than imported, since these two excel/ scripts are each
    meant to stand alone."""
    candidates = []
    for f in CLEAN_DIR.glob("week_*_predictions.csv"):
        m = PRED_FILE_RE.match(f.name)
        if m:
            candidates.append(((int(m.group(1)), int(m.group(2))), f))
    if not candidates:
        return None
    candidates.sort(key=lambda t: t[0])
    return candidates[-1][1]


def _lean(model_spread_home, market_spread_home):
    """Same edge/lean convention as backtest.py's/model_comparison.py's own
    _grade_side() -- edge = market minus model, sign says which side the
    model likes relative to the market. None (not "Pick'em") when there's no
    market line yet to compare against, since that's a different situation
    from a real, computed pick'em."""
    if market_spread_home is None or pd.isna(market_spread_home):
        return None
    edge = market_spread_home - model_spread_home
    if edge > 0:
        return "Home"
    if edge < 0:
        return "Away"
    return "Pick'em"


def write_upcoming_block(ws, start_row, con):
    """
    Ridge's and XGBoost's live picks for the next upcoming week, side by
    side -- NOT graded (the games haven't been played yet), just the two
    models' current lines plus whatever market line has posted so far.
    Reads predict_week.py's latest predictions CSV, which now carries an
    'XGBoost Line (Home)' column alongside Ridge's (see that script's
    docstring for why this doesn't touch Weekly Slate/Ridge's live-model
    status at all).

    "Tighter (Live)*" is a PROVISIONAL version of the graded table's
    "Tighter CLV Line" column below -- it compares each model's line to
    TODAY's still-moving market line, not a real closing line (the game
    hasn't kicked off, so there isn't one yet). It'll keep changing as the
    market moves and is only meaningful as a snapshot of "as of right now" --
    the asterisk plus the note row under the table exist so this is never
    mistaken for a final CLV read the way the graded table's version is.
    """
    r = start_row
    ws.cell(row=r, column=1, value="Upcoming Games -- Ridge vs. XGBoost (Not Yet Graded)").font = SECTION_FONT
    r += 1

    pred_path = latest_predictions_file()
    if pred_path is None:
        ws.cell(row=r, column=1,
                value="No predictions CSV found yet -- run src/predict_week.py first.").font = NOTE_FONT
        return r + 2, None

    preds = pd.read_csv(pred_path)
    if preds.empty or "XGBoost Line (Home)" not in preds.columns:
        ws.cell(row=r, column=1,
                value="Latest predictions CSV has no XGBoost column yet -- rerun src/predict_week.py "
                      "(it now produces one).").font = NOTE_FONT
        return r + 2, None

    game_ids = [int(g) for g in preds["Game ID"].tolist()]
    placeholders = ",".join("?" * len(game_ids))
    market = dict(con.execute(f"""
        SELECT game_id, AVG(spread) FROM lines WHERE game_id IN ({placeholders}) GROUP BY game_id
    """, game_ids).fetchall()) if game_ids else {}

    headers = ["Week", "Matchup", "Market Line (Home)", "Ridge Line", "XGBoost Line",
               "Ridge Lean", "XGBoost Lean", "Models Agree?", "Tighter (Live)*"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    header_row = r
    r += 1

    def fmt_spread(v):
        if v is None or pd.isna(v):
            return "--"
        return f"+{v:.1f}" if v > 0 else f"{v:.1f}"

    # .to_dict("records") + plain string column access, not itertuples() --
    # several of these column names have spaces ("Away Team", "Model Line
    # (Home)"), which itertuples() mangles into positional _N attributes
    # that are too easy to miscount and silently read the wrong column.
    for row in preds.to_dict("records"):
        matchup = f"{row['Away Team']} @ {row['Home Team']}"
        market_line = market.get(int(row["Game ID"]))
        ridge_line = row["Model Line (Home)"]
        xgb_line = row["XGBoost Line (Home)"]
        ridge_lean = _lean(ridge_line, market_line)
        xgb_lean = _lean(xgb_line, market_line)
        agree = None
        if ridge_lean not in (None, "Pick'em") and xgb_lean not in (None, "Pick'em"):
            agree = ridge_lean == xgb_lean

        if market_line is None or pd.isna(market_line):
            tighter_live = "n/a"
        else:
            ridge_abs, xgb_abs = abs(ridge_line - market_line), abs(xgb_line - market_line)
            tighter_live = "Ridge" if ridge_abs < xgb_abs else ("XGBoost" if xgb_abs < ridge_abs else "Tie")

        values = [
            int(row["Week"]), matchup, fmt_spread(market_line),
            fmt_spread(ridge_line), fmt_spread(xgb_line),
            ridge_lean or "n/a", xgb_lean or "n/a",
            ("Yes" if agree is True else ("No" if agree is False else "n/a")),
            tighter_live,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = FORMULA_FONT
            if c == 9 and tighter_live in ("Ridge", "XGBoost"):
                cell.fill = WIN_FILL
        r += 1

    ws.cell(row=r, column=1,
            value="* Tighter (Live) compares each model's line to TODAY's still-moving market line, not a final "
                  "closing line -- it will keep changing until kickoff. See the graded Game-by-Game table below "
                  "for the real, closing-line version of this comparison.").font = NOTE_FONT
    r += 1

    return r + 1, header_row


def build_sheet(wb, df, con):
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Ridge vs. XGBoost -- Model Comparison"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Informational only. Ridge is still the live model everywhere else in this workbook and on "
        "the website -- XGBoost is a candidate being evaluated here, not a replacement."
    )
    ws["A2"].font = NOTE_FONT

    current_season = int(df["season"].max())

    row = 4
    row = write_summary_block(ws, df, row, "All-Time (every season in the backtest)")
    row = write_summary_block(ws, df, row, f"{current_season} Season Only", season_filter=current_season)

    agree_rate = df["models_agree"].dropna().mean() if df["models_agree"].notna().any() else None
    if agree_rate is not None:
        ws.cell(row=row, column=1,
                value=f"Ridge and XGBoost agree on which side to lean in {agree_rate:.1%} of all graded games.").font = NOTE_FONT
        row += 2

    row, _ = write_upcoming_block(ws, row, con)
    row += 1

    row, header_row = write_detail_table(ws, df, row, current_season)

    # 17 columns now -- widened where a column's meaning differs between the
    # stacked sub-tables that share it (e.g. col 9 is "Ridge Result" in the
    # graded table but "Tighter Than Other (CLV)" in the summary blocks).
    autosize(ws, [8, 26, 12, 18, 18, 12, 12, 11, 18, 16, 12, 11, 12, 13, 15, 17, 17])
    # NOT header_row-based: this sheet stacks several tables (summary blocks,
    # the upcoming-games table, the graded detail table), each with its own
    # header further down the sheet -- freezing through the last one (as an
    # earlier version of this function did) pins the ENTIRE top of the sheet
    # in place, which on a screen that can't fit that many rows leaves no way
    # to scroll the frozen region into view at all. Freeze just the title/
    # subtitle (rows 1-2) instead, same as a simple top-freeze on any other
    # tab -- everything below scrolls normally.
    ws.freeze_panes = "A3"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    if not RESULTS_PATH.exists():
        print(f"{RESULTS_PATH} not found. Run 'python src/model_comparison.py' first.")
        return
    if not TRACKER_PATH.exists():
        print(f"Tracker workbook not found at {TRACKER_PATH}. Run excel/build_tracker.py first.")
        return

    df = pd.read_csv(RESULTS_PATH)
    if df.empty:
        print("model_comparison_results.csv is empty -- nothing to write yet.")
        return
    # models_agree round-trips through CSV as True/False/NaN -- pandas already
    # reads that correctly as object dtype with real bool/NaN values.

    wb = openpyxl.load_workbook(TRACKER_PATH)
    con = duckdb.connect(str(DB_PATH))
    build_sheet(wb, df, con)
    con.close()
    wb.save(TRACKER_PATH)
    print(f"'{SHEET_NAME}' tab written to {TRACKER_PATH}. Every other tab is untouched.")


if __name__ == "__main__":
    main()
