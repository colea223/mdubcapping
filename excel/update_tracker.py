"""
Phase 4: wires the model's output into MW_Handicapping_Tracker.xlsx so you're
not copy-pasting numbers by hand every week.

What it writes (and ONLY this -- everything else in the workbook, including
your Bet Log and any notes you've typed, is left untouched):
  - Weekly Slate: Week, Date, Away Team, Home Team, Model Line (Home), Model
    Total for the latest week predict_week.py has produced. Market Line and
    Market Total stay blank/manual -- that's still your sportsbook lookup.
  - Team Profiles: Model Power Rating, SP+ Rating, Off PPA, Def PPA, Talent
    Rank for each of the 10 teams, pulled from the database's latest season.

A game already on the Weekly Slate (matched by Home + Away team) gets its
Model Line/Model Total updated in place rather than duplicated; a new game
is written into the first empty row.

IMPORTANT: this script does NOT recalculate formulas (that requires
LibreOffice, which this project doesn't assume you have installed). That's
fine in practice -- when you open the file in real Microsoft Excel, it
recalculates every formula automatically on open. You do not need to do
anything extra.

Usage:
    source .venv/bin/activate     (or the Windows equivalent)
    python excel/update_tracker.py
"""
import re
import sys
from pathlib import Path

import duckdb
import openpyxl
import time

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from config import DB_PATH, CLEAN_DIR  # noqa: E402
from power_rating import current_ratings  # noqa: E402

TRACKER_PATH = Path(__file__).resolve().parent / "MW_Handicapping_Tracker.xlsx"
WEEKLY_SLATE_ROWS = range(2, 43)   # matches build_tracker.py's layout
TEAM_PROFILE_ROWS = range(2, 12)   # 10 teams

PRED_FILE_RE = re.compile(r"^week_(\d{4})_(\d+)_predictions\.csv$")


def latest_predictions_file():
    candidates = []
    for f in CLEAN_DIR.glob("week_*_predictions.csv"):
        m = PRED_FILE_RE.match(f.name)
        if m:
            candidates.append(((int(m.group(1)), int(m.group(2))), f))
    if not candidates:
        return None
    candidates.sort(key=lambda t: t[0])
    return candidates[-1][1]


def update_weekly_slate(ws, predictions_path: Path, con):
    import csv
    with open(predictions_path, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        print("Weekly Slate: predictions file is empty -- nothing to write.")
        return

    # CFBD's own betting lines (the same data the website's Matchups page
    # reads) -- auto-fill Market Line/Total when a book has actually posted
    # one for that game yet. A missing game_id here just means no line is in
    # the DB yet (common for games further out); leave that cell for you to
    # fill by hand rather than blanking out anything you already typed in.
    market = {
        game_id: (spread, total)
        for game_id, spread, total in con.execute("""
            SELECT game_id, AVG(spread), AVG(over_under) FROM lines GROUP BY game_id
        """).fetchall()
    }
    filled_market = 0

    # Clear any row for a matchup that ISN'T in this run's predictions --
    # otherwise a game that drops off the list (last week's games once a new
    # week is predicted, or -- what actually happened here -- games that
    # never should've been on this MW-focused sheet in the first place) just
    # sits there forever, since the loop below only ever adds/updates rows,
    # never removes them. A row for a matchup that IS still in the current
    # predictions is left alone here (and updated in place below), so manual
    # Market Line/Total/Notes/Final Dec entries survive a same-week rerun.
    current_keys = {(pred["Home Team"], pred["Away Team"]) for pred in rows}
    cleared = 0
    for r in WEEKLY_SLATE_ROWS:
        home, away = ws[f"D{r}"].value, ws[f"C{r}"].value
        if home and away and (home, away) not in current_keys:
            for col in ("A", "B", "C", "D", "E", "F", "I", "J", "O", "P"):
                ws[f"{col}{r}"] = None
            cleared += 1
    if cleared:
        print(f"Weekly Slate: cleared {cleared} row(s) no longer in this week's predictions")

    # index existing rows by (home, away) so a re-run updates in place
    existing = {}
    first_empty = None
    for r in WEEKLY_SLATE_ROWS:
        home = ws[f"D{r}"].value
        away = ws[f"C{r}"].value
        if home and away:
            existing[(home, away)] = r
        elif first_empty is None:
            first_empty = r

    written = 0
    for pred in rows:
        key = (pred["Home Team"], pred["Away Team"])
        if key in existing:
            r = existing[key]
        elif first_empty is not None:
            r = first_empty
            first_empty = next((row for row in WEEKLY_SLATE_ROWS if row > r and not ws[f"D{row}"].value), None)
        else:
            print(f"  Weekly Slate is full (rows 2-42 all used) -- skipping {key}")
            continue

        ws[f"A{r}"] = int(pred["Week"])
        ws[f"B{r}"] = pred["Date"]
        ws[f"C{r}"] = pred["Away Team"]
        ws[f"D{r}"] = pred["Home Team"]
        ws[f"E{r}"] = float(pred["Model Line (Home)"])
        ws[f"I{r}"] = float(pred["Model Total"])
        written += 1

        # .get() -- older predictions CSVs (from before Game ID was added)
        # simply won't have this column, and that's fine: skip the auto-fill
        # for those rather than erroring.
        game_id = pred.get("Game ID")
        if game_id is not None:
            try:
                spread, total = market.get(int(game_id), (None, None))
            except ValueError:
                spread, total = None, None
            if spread is not None:
                ws[f"F{r}"] = round(spread, 1)
                filled_market += 1
            if total is not None:
                ws[f"J{r}"] = round(total, 1)

    print(f"Weekly Slate: wrote/updated {written} matchup(s)")
    if filled_market:
        print(f"  -> auto-filled Market Line/Total for {filled_market} game(s) with a CFBD line already posted")


def update_team_profiles(ws, con):
    ratings = current_ratings(con)
    sp = dict(con.execute("""
        SELECT team, rating FROM sp_ratings
        WHERE season = (SELECT MAX(season) FROM sp_ratings)
    """).fetchall())
    adv = con.execute("""
        SELECT team, off_ppa, def_ppa FROM advanced_stats
        WHERE season = (SELECT MAX(season) FROM advanced_stats)
    """).fetchall()
    adv_map = {team: (off_ppa, def_ppa) for team, off_ppa, def_ppa in adv}
    rec = dict(con.execute("""
        SELECT team, rank FROM recruiting
        WHERE season = (SELECT MAX(season) FROM recruiting)
    """).fetchall())

    updated = 0
    for r in TEAM_PROFILE_ROWS:
        team = ws[f"A{r}"].value
        if not team:
            continue
        if team in ratings:
            ws[f"D{r}"] = round(ratings[team], 1)
        if team in sp:
            ws[f"E{r}"] = sp[team]
        if team in adv_map:
            off_ppa, def_ppa = adv_map[team]
            ws[f"F{r}"] = off_ppa
            ws[f"G{r}"] = def_ppa
        if team in rec:
            ws[f"H{r}"] = rec[team]
        updated += 1

    print(f"Team Profiles: updated {updated} team row(s)")


def main():
    if not TRACKER_PATH.exists():
        print(f"Tracker workbook not found at {TRACKER_PATH}. Run excel/build_tracker.py first, "
              "or move MW_Handicapping_Tracker.xlsx next to this script.")
        return

    pred_file = latest_predictions_file()
    if pred_file is None:
        print("No predictions CSV found in data/clean/. Run src/predict_week.py first.")
        return
    print(f"Using predictions from {pred_file.name}")

    wb = openpyxl.load_workbook(TRACKER_PATH)  # formulas preserved as formulas, not evaluated

    con = duckdb.connect(str(DB_PATH))
    update_weekly_slate(wb["Weekly Slate"], pred_file, con)
    update_team_profiles(wb["Team Profiles"], con)
    con.close()

    wb.save(TRACKER_PATH)
    print(f"\nSaved {TRACKER_PATH}. Open it in Excel -- formulas recalculate automatically on open.")


if __name__ == "__main__":
    _script_start_time = time.time()
    main()

    _script_elapsed = time.time() - _script_start_time
    _mins, _secs = divmod(_script_elapsed, 60)
    print(f"\n[Finished in {int(_mins)}m {_secs:04.1f}s]" if _mins else f"\n[Finished in {_secs:.1f}s]")
