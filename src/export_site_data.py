"""
Produces the JSON files the website (docs/) reads: rankings, this week's
matchups (market lines only, Covers-style), predictions (model output + any
manually-noted injury/qualitative info), and season-to-date tracking stats.

Injuries/qualitative notes have no reliable free automated feed (CFBD doesn't
publish one), so that piece is manually maintained: edit site_notes.json in
the project root -- {"Away Team @ Home Team": "note text"} -- and this script
merges it in by matchup. Leave it out or empty and notes just render blank.

The Tracking page shows THREE things side by side: the live model's own
hypothetical flat-1-unit-stake performance (same walk-forward grading as
backtest.py, filtered to the current season -- spread, total, AND moneyline),
YOUR actual bets, read directly from the Bet Log tab of
excel/MW_Handicapping_Tracker.xlsx, and the "Dub Beta Model" (XGBoost, spread
only) -- see below. None of these are ever blended into one number -- the
live model's picks are a what-if, your Bet Log is what you actually staked,
and the Dub Beta Model is a candidate architecture being evaluated alongside
the live model, not a replacement for it.

Dub Beta Model (XGBoost) placement, and why it's sourced the way it is:
  - Predictions page: this week's Dub Beta line comes from predict_week.py's
    latest week_<season>_<week>_predictions.csv ("XGBoost Line (Home)"
    column) -- NOT refit here. predict_week.py already runs as its own
    pipeline step before this one (see run_pipeline.py's STEPS), so reading
    its output avoids paying for the same XGBoost hyperparameter search
    twice in one pipeline run.
  - Results page & season tracking: graded Dub Beta history comes from
    data/clean/model_comparison_results.csv, written by src/model_comparison.py.
    That script is deliberately a standalone, hand-run script (same category
    as backtest.py) because it reruns the XGBoost search once per historical
    test week -- wiring it into every site export would make a routine
    pipeline run as slow as a full backtest. Practically: this means the Dub
    Beta Model's Results/Tracking numbers are only as fresh as the last time
    you ran `python src/model_comparison.py` by hand, same staleness contract
    excel/update_model_comparison_tab.py already has with this file. And
    since model_comparison.py compares margin models only, Dub Beta's graded
    record is spread-only -- it has no total/moneyline pick to show.

Usage:
    source .venv/bin/activate
    python src/export_site_data.py
"""
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

from config import DB_PATH, CLEAN_DIR
import model
import totals_model
import backtest
from odds import payout_profit
from power_rating import current_ratings
from teams import MW_TEAMS_2026
from predict_week import auto_detect_week
from features import team_home_venues, haversine_km

ROOT = Path(__file__).resolve().parent.parent
DOCS_DATA = ROOT / "docs" / "data"
NOTES_PATH = ROOT / "site_notes.json"
MANUAL_LINES_PATH = ROOT / "manual_lines.json"
TRACKER_PATH = ROOT / "excel" / "MW_Handicapping_Tracker.xlsx"
BET_LOG_ROWS = range(2, 43)   # matches excel/build_tracker.py's layout
CURRENT_SEASON = 2026

# Dub Beta Model (XGBoost) data sources -- see module docstring above.
PRED_FILE_RE = re.compile(r"^week_(\d{4})_(\d+)_predictions\.csv$")
MODEL_COMPARISON_PATH = CLEAN_DIR / "model_comparison_results.csv"


def latest_predictions_file():
    """Same convention/name as excel/update_model_comparison_tab.py's own
    helper (itself duplicated from excel/update_tracker.py's) -- duplicated
    again here rather than imported, since every script in this project is
    meant to stand alone. Finds predict_week.py's most recent
    week_<season>_<week>_predictions.csv."""
    candidates = []
    for f in CLEAN_DIR.glob("week_*_predictions.csv"):
        m = PRED_FILE_RE.match(f.name)
        if m:
            candidates.append(((int(m.group(1)), int(m.group(2))), f))
    if not candidates:
        return None
    candidates.sort(key=lambda t: t[0])
    return candidates[-1][1]


def load_beta_comparison():
    """
    Dub Beta Model's graded history -- data/clean/model_comparison_results.csv,
    written by src/model_comparison.py's walk-forward Ridge-vs-XGBoost
    comparison. Returns None if that file doesn't exist yet (script never
    run) or is empty, so every caller can just treat "no Dub Beta data" as a
    normal, expected state -- same graceful-degradation pattern as
    load_manual_lines()/load_notes() elsewhere in this file.
    """
    if not MODEL_COMPARISON_PATH.exists():
        return None
    df = pd.read_csv(MODEL_COMPARISON_PATH)
    return df if not df.empty else None


def _now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _sanitize_nans(obj):
    """
    Recursively replaces float NaN with None (JSON null), everywhere in a
    dict/list structure, right before it's written out.

    Root cause: build_results() (and potentially other builders here) pulls
    fields straight out of a pandas DataFrame built from a list of dicts
    where a not-applicable value was Python None (e.g. backtest.py's
    bet_result/total_bet_result/ml_bet_result default to None). When a
    DataFrame column mixes strings and None, pandas silently upcasts those
    Nones to float NaN -- so accessing the value back out via itertuples()
    returns nan, not None. Python's json.dumps() then happily writes that as
    a bare, UNQUOTED "NaN" token, which is not valid JSON (the spec has no
    such literal) -- browsers' JSON.parse() throws a hard SyntaxError on it,
    which silently breaks the whole file, not just that one field. This is
    the same None-vs-NaN trap documented in excel/update_model_comparison_tab.py's
    result_text() helper, just showing up here via a different code path.
    Sanitizing the whole structure right before every json.dumps() call
    below is more robust than chasing down each individual field that could
    carry a None through a DataFrame.
    """
    if isinstance(obj, float) and math.isnan(obj):
        return None
    if isinstance(obj, dict):
        return {k: _sanitize_nans(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_nans(v) for v in obj]
    return obj


def load_notes():
    if NOTES_PATH.exists():
        return json.loads(NOTES_PATH.read_text())
    return {}


def load_manual_lines():
    """
    Hand-typed fallback lines for when the pipeline can't reach CFBD (API
    limit hit, an outage, whatever) and a game would otherwise show up on
    the site with no market number at all. Edit manual_lines.json directly
    -- same "Away Team @ Home Team" key as site_notes.json -- with any
    subset of: spread_home (negative = home favored), total, home_ml,
    away_ml. This is DISPLAY ONLY: it's never written into the lines table
    itself, never touches the model or backtest, and it only ever fills in
    a number that would otherwise be blank -- it can't override a real
    pulled line. Leave the file out or empty and nothing changes.
    """
    if MANUAL_LINES_PATH.exists():
        return json.loads(MANUAL_LINES_PATH.read_text())
    return {}


def team_record(con, team, season):
    row = con.execute("""
        SELECT
            SUM(CASE WHEN home_team = ? AND home_points > away_points THEN 1
                     WHEN away_team = ? AND away_points > home_points THEN 1 ELSE 0 END) AS wins,
            SUM(CASE WHEN (home_team = ? AND home_points < away_points)
                       OR (away_team = ? AND away_points < home_points) THEN 1 ELSE 0 END) AS losses
        FROM games
        WHERE completed = TRUE AND (home_team = ? OR away_team = ?) AND season = ?
    """, [team, team, team, team, team, team, season]).fetchone()
    wins, losses = row[0] or 0, row[1] or 0
    return wins, losses


def team_conference_record(con, team, season):
    """
    Same shape as team_record(), restricted to CFBD's own conference_game
    flag -- games where both teams belonged to the same conference that
    season, per CFBD's own classification (not this project's MW_TEAMS_2026
    team-name membership). For 2026 that correctly counts only true
    Mountain West-vs-Mountain West games for every current member, including
    UTEP and Northern Illinois now that they've actually joined -- no
    separate newcomer-handling needed here since conference_game already
    reflects each season's real alignment.
    """
    row = con.execute("""
        SELECT
            SUM(CASE WHEN home_team = ? AND home_points > away_points THEN 1
                     WHEN away_team = ? AND away_points > home_points THEN 1 ELSE 0 END) AS wins,
            SUM(CASE WHEN (home_team = ? AND home_points < away_points)
                       OR (away_team = ? AND away_points < home_points) THEN 1 ELSE 0 END) AS losses
        FROM games
        WHERE completed = TRUE AND conference_game = TRUE
          AND (home_team = ? OR away_team = ?) AND season = ?
    """, [team, team, team, team, team, team, season]).fetchone()
    wins, losses = row[0] or 0, row[1] or 0
    return wins, losses


def team_rating_trend(con, team):
    rows = con.execute("""
        SELECT r.rating_after, g.start_date
        FROM ratings_baseline r
        JOIN games g ON g.game_id = r.game_id
        WHERE r.team = ?
        ORDER BY g.start_date DESC
        LIMIT 2
    """, [team]).fetchall()
    if len(rows) < 2:
        return 0.0
    return round(rows[0][0] - rows[1][0], 1)


def build_rankings(con):
    ratings = current_ratings(con)
    rows = []
    for team in MW_TEAMS_2026:
        rating = ratings.get(team)
        wins, losses = team_record(con, team, CURRENT_SEASON)
        conf_wins, conf_losses = team_conference_record(con, team, CURRENT_SEASON)
        trend = team_rating_trend(con, team)
        rows.append({
            "team": team,
            "rating": round(rating, 1) if rating is not None else None,
            "wins": wins, "losses": losses,
            "conf_wins": conf_wins, "conf_losses": conf_losses,
            "trend": trend,
        })
    rows.sort(key=lambda r: (r["rating"] is None, -(r["rating"] or 0)))
    for i, r in enumerate(rows, start=1):
        r["rank"] = i
    return rows


def mw_game(row_home, row_away):
    return row_home in MW_TEAMS_2026 or row_away in MW_TEAMS_2026


# Providers in CFBD's lines table that are NOT real sportsbooks -- analytics
# sites whose "lines" are power-rating projections, not bettable market
# prices. Excluded everywhere on the live-lines page so nothing there looks
# like a book you could actually place a wager with.
NOT_REAL_BOOKS = {"teamrankings", "numberfire"}

# The three sportsbooks featured as the Live Lines table's main book columns
# (Cole's picks: DraftKings, FanDuel, Bovada). Bet365 was considered but
# isn't offered for NCAAF through The Odds API at all -- confirmed against
# their own bookmaker-region docs (Bet365 is listed only under the Australia
# region, for AFL/NRL) and against this project's own historical pulls,
# which have never once returned it -- so it was dropped rather than left in
# as a column that would always read "no line." Matched against
# build_db.py's normalize_provider() output, so "Draft Kings"/"DraftKings"
# variants both collapse to the one name here. "consensus" (CFBD's own
# blended-line row, not an actual book) is kept alongside these three
# regardless -- it's a reference point, not one of "the three main books."
MAIN_BOOKS = {"DraftKings", "FanDuel", "Bovada"}


def build_live_lines(con, manual_lines=None):
    """
    A per-book live line table for the current MW slate -- open line plus
    every real sportsbook's current spread/total/moneyline, so you can see
    both the current number and how far it's moved from open, book by book.
    Mirrors the covers.com-style line tracker: Time / Game / Open / [books].

    Also computes a CLOSING line once a game's own kickoff has passed --
    auto_detect_week() keeps returning the same season/week as long as ANY
    game that week is still incomplete, so an early-week game (a Thursday
    MW game, say) sits on the same page as the rest of that week's still-
    upcoming games. Until kickoff, "close_*" is left null (there's no closing
    number yet -- the book columns are still live); once kickoff passes, it's
    the same real-books consensus average the "current" book columns already
    show, just labeled and frozen as the closing reference point rather than
    something that reads as still-moving. This mirrors exactly what
    backtest.py treats as market_close for a graded game -- the same lines
    table columns, just surfaced here for the live-lines page too, per-game,
    the moment kickoff passes rather than only after backtest.py runs.

    manual_lines (see load_manual_lines()) adds one extra "Manual" column
    for any matchup listed there -- purely a display fallback for when the
    pipeline couldn't pull real lines, never a replacement for real ones.
    """
    manual_lines = manual_lines or {}
    now = datetime.now(timezone.utc)
    detected = auto_detect_week(con)
    if detected is None:
        return [], None
    season, week = detected[0], detected[1]

    games = con.execute("""
        SELECT game_id, season, week, start_date, home_team, away_team
        FROM games WHERE season = ? AND week = ?
        ORDER BY start_date
    """, [season, week]).fetchdf()
    games = games[games.apply(lambda r: mw_game(r["home_team"], r["away_team"]), axis=1)]
    if games.empty:
        return [], {"season": season, "week": week}

    game_ids = games["game_id"].tolist()
    placeholders = ",".join("?" * len(game_ids))
    # Restrict book columns to MAIN_BOOKS (+ consensus) -- see that constant's
    # comment for why Bet365 isn't one of them. NOT_REAL_BOOKS is still
    # excluded explicitly too, purely defensive in case MAIN_BOOKS ever grows
    # to include something that overlaps it.
    allowed_books = list(MAIN_BOOKS | {"consensus"})
    lines = con.execute(f"""
        SELECT game_id, provider, spread, spread_open, over_under, over_under_open,
               home_moneyline, away_moneyline
        FROM lines
        WHERE game_id IN ({placeholders})
          AND provider IN ({",".join("?" * len(allowed_books))})
          AND provider NOT IN ({",".join("?" * len(NOT_REAL_BOOKS))})
    """, game_ids + allowed_books + list(NOT_REAL_BOOKS)).fetchdf()

    # CFBD's own `lines` endpoint (the query above) never returns FanDuel at
    # all -- confirmed against this project's real pulls, it only ever hands
    # back Bovada/DraftKings. FanDuel is one of MAIN_BOOKS anyway (Cole's
    # pick), and it DOES show up through the second, independent feed --
    # The Odds API (pull_odds_api.py) -- which lands in line_snapshots
    # tagged source='odds_api'. That table is otherwise only used for the
    # Line History chart; this pulls in just the latest odds_api reading per
    # (game, provider) as a supplemental "current" column for whichever
    # MAIN_BOOKS provider CFBD didn't cover for that game. It intentionally
    # does NOT feed open_spread_home/open_total/close_spread_home/
    # close_total above -- those stay CFBD-`lines`-only so they keep
    # matching backtest.py's market_close exactly. And it never overrides a
    # book CFBD DID return (Bovada/DraftKings keep their true CFBD
    # spread_open/over_under_open); this only fills in a gap.
    odds_api_books = con.execute(f"""
        SELECT game_id, provider, spread, over_under, home_moneyline, away_moneyline
        FROM line_snapshots
        WHERE game_id IN ({placeholders})
          AND source = 'odds_api'
          AND provider IN ({",".join("?" * len(MAIN_BOOKS))})
        QUALIFY ROW_NUMBER() OVER (PARTITION BY game_id, provider ORDER BY pulled_at DESC) = 1
    """, game_ids + list(MAIN_BOOKS)).fetchdf()

    rows = []
    for g in games.itertuples():
        g_lines = lines[lines["game_id"] == g.game_id]
        real_books = g_lines[g_lines["provider"] != "consensus"]
        open_source = real_books if not real_books.empty else g_lines

        def _avg(col, source=open_source):
            vals = source[col].dropna()
            return round(float(vals.mean()), 1) if not vals.empty else None

        kicked_off = bool(pd.to_datetime(g.start_date, utc=True) <= now)
        # "current" spread/over_under from real books -- once kickoff has
        # passed these stop moving (nothing new gets pulled for a game
        # that's already underway/final), so the same consensus average is
        # both "the latest number" and "the closing number" at that point.
        close_source = real_books if not real_books.empty else g_lines
        close_spread_home = _avg("spread", close_source) if kicked_off else None
        close_total = _avg("over_under", close_source) if kicked_off else None

        books = {}
        for r in g_lines.itertuples():
            books[r.provider] = {
                "spread": round(r.spread, 1) if pd.notna(r.spread) else None,
                "spread_open": round(r.spread_open, 1) if pd.notna(r.spread_open) else None,
                "total": round(r.over_under, 1) if pd.notna(r.over_under) else None,
                "total_open": round(r.over_under_open, 1) if pd.notna(r.over_under_open) else None,
                "home_ml": int(r.home_moneyline) if pd.notna(r.home_moneyline) else None,
                "away_ml": int(r.away_moneyline) if pd.notna(r.away_moneyline) else None,
            }
        # Fill in FanDuel (or any other MAIN_BOOKS provider CFBD didn't
        # return for this game) from The Odds API's latest reading. No
        # spread_open/total_open here -- The Odds API has no "open" concept
        # of its own and CFBD never carried this provider to compare
        # against, so leaving those null is the honest answer rather than
        # faking a movement indicator, same treatment build_line_history()
        # already gives a provider missing from the `lines` table.
        g_odds_api = odds_api_books[odds_api_books["game_id"] == g.game_id]
        for r in g_odds_api.itertuples():
            if r.provider in books:
                continue
            books[r.provider] = {
                "spread": round(r.spread, 1) if pd.notna(r.spread) else None,
                "spread_open": None,
                "total": round(r.over_under, 1) if pd.notna(r.over_under) else None,
                "total_open": None,
                "home_ml": int(r.home_moneyline) if pd.notna(r.home_moneyline) else None,
                "away_ml": int(r.away_moneyline) if pd.notna(r.away_moneyline) else None,
            }
        manual = manual_lines.get(f"{g.away_team} @ {g.home_team}")
        if manual:
            books["Manual"] = {
                "spread": manual.get("spread_home"), "spread_open": None,
                "total": manual.get("total"), "total_open": None,
                "home_ml": manual.get("home_ml"), "away_ml": manual.get("away_ml"),
            }

        # Stable order: Consensus first (if present), real books alphabetically,
        # Manual last so it never gets mistaken for an actual sportsbook price.
        book_order = sorted(
            books.keys(),
            key=lambda p: (p != "consensus", p == "Manual", p.lower()),
        )

        rows.append({
            "game_id": int(g.game_id),
            "date": pd.to_datetime(g.start_date).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "away_team": g.away_team,
            "home_team": g.home_team,
            "open_spread_home": _avg("spread_open"),
            "open_total": _avg("over_under_open"),
            "kicked_off": kicked_off,
            "close_spread_home": close_spread_home,
            "close_total": close_total,
            "books": books,
            "book_order": book_order,
        })

    return rows, {"season": season, "week": week}


def build_line_history(con, game_ids):
    """
    Every historical snapshot (line_snapshots -- see build_db.py) for the
    given games, shaped for a per-game/per-book/per-market chart: a time
    series the Live Lines page can plot, plus the true open value as a fixed
    reference point. Early on (or for a game just added) this may be as few
    as one point -- it only gets richer as pull_lines.py runs more often
    over the season, since nothing here is retroactive.
    """
    if not game_ids:
        return {}
    placeholders = ",".join("?" * len(game_ids))
    snaps = con.execute(f"""
        SELECT game_id, provider, pulled_at, spread, over_under, home_moneyline, away_moneyline, source
        FROM line_snapshots
        WHERE game_id IN ({placeholders})
        ORDER BY pulled_at
    """, game_ids).fetchdf()
    opens = con.execute(f"""
        SELECT game_id, provider, spread_open, over_under_open
        FROM lines
        WHERE game_id IN ({placeholders})
    """, game_ids).fetchdf().set_index(["game_id", "provider"])

    history = {}
    for gid, g in snaps.groupby("game_id"):
        by_provider = {}
        for provider, p in g.groupby("provider"):
            open_row = opens.loc[(gid, provider)] if (gid, provider) in opens.index else None
            by_provider[provider] = {
                "open_spread": round(float(open_row["spread_open"]), 1) if open_row is not None and pd.notna(open_row["spread_open"]) else None,
                "open_total": round(float(open_row["over_under_open"]), 1) if open_row is not None and pd.notna(open_row["over_under_open"]) else None,
                "points": [
                    {
                        "t": row.pulled_at.strftime("%Y-%m-%dT%H:%M:%SZ"),
                        "spread": round(row.spread, 1) if pd.notna(row.spread) else None,
                        "total": round(row.over_under, 1) if pd.notna(row.over_under) else None,
                        "home_ml": int(row.home_moneyline) if pd.notna(row.home_moneyline) else None,
                        "away_ml": int(row.away_moneyline) if pd.notna(row.away_moneyline) else None,
                        "source": row.source if pd.notna(row.source) else "cfbd",
                    }
                    for row in p.itertuples()
                ],
            }
        history[str(int(gid))] = by_provider
    return history


def build_matchups_and_predictions(con, notes: dict, manual_lines=None):
    manual_lines = manual_lines or {}
    detected = auto_detect_week(con)
    if detected is None:
        return [], [], None
    season, week = detected[0], detected[1]

    games = con.execute("""
        SELECT game_id, season, week, start_date, home_team, away_team,
               completed, home_points, away_points
        FROM games WHERE season = ? AND week = ?
        ORDER BY start_date
    """, [season, week]).fetchdf()
    games = games[games.apply(lambda r: mw_game(r["home_team"], r["away_team"]), axis=1)]

    lines = con.execute("""
        SELECT game_id, AVG(spread) AS spread, AVG(over_under) AS total,
               AVG(home_moneyline) AS home_ml, AVG(away_moneyline) AS away_ml
        FROM lines GROUP BY game_id
    """).fetchdf().set_index("game_id")

    matchups = []
    for g in games.itertuples():
        line = lines.loc[g.game_id] if g.game_id in lines.index else None
        manual = manual_lines.get(f"{g.away_team} @ {g.home_team}") or {}

        def pick(real_val, manual_key, round_to=None):
            # Real pulled data always wins; manual_lines.json only fills a
            # gap left by a failed/missing pull -- see load_manual_lines().
            if real_val is not None and pd.notna(real_val):
                return round(real_val, round_to) if round_to is not None else round(real_val)
            mv = manual.get(manual_key)
            return round(mv, round_to) if (mv is not None and round_to is not None) else mv

        # A week runs Thu-Sat/Sun, so auto_detect_week() keeps returning this
        # same (season, week) as long as ANY game in it is still incomplete --
        # meaning an earlier game in the same week can already be final while
        # this page still treats the week as "current." completed/home_points/
        # away_points let the front end show a Final score for those rather
        # than leaving them looking perpetually upcoming.
        completed = bool(g.completed)
        home_pts = int(g.home_points) if completed and pd.notna(g.home_points) else None
        away_pts = int(g.away_points) if completed and pd.notna(g.away_points) else None

        matchups.append({
            "game_id": int(g.game_id), "week": int(g.week),
            "date": pd.to_datetime(g.start_date).strftime("%Y-%m-%d"),
            "away_team": g.away_team, "home_team": g.home_team,
            "market_spread_home": pick(line["spread"] if line is not None else None, "spread_home", 1),
            "market_total": pick(line["total"] if line is not None else None, "total", 1),
            "home_moneyline": pick(line["home_ml"] if line is not None else None, "home_ml"),
            "away_moneyline": pick(line["away_ml"] if line is not None else None, "away_ml"),
            "manual_line_used": bool(manual and (line is None or bool(line.isna().all()))),
            "completed": completed,
            "home_points": home_pts,
            "away_points": away_pts,
        })

    predictions = []
    train_df = model.load_training_frame(con)
    print(f"  [predictions diag] train_df: {len(train_df)} completed games w/ a market line "
          f"(need >= 10) -- games this week (MW-involved): {len(games)}")
    if len(train_df) >= 10 and not games.empty:
        pipe, residual_std = model.fit_margin_model(train_df)
        upcoming = model.load_upcoming_frame(con, season, week)
        upcoming_before = len(upcoming)
        upcoming = upcoming[upcoming["game_id"].isin(games["game_id"])]
        print(f"  [predictions diag] game_features rows for week {week}: {upcoming_before} "
              f"-- matching this week's MW games: {len(upcoming)}")
        if not upcoming.empty:
            pred_margin = model.predict_margin(pipe, upcoming)
            model_spread_home = -pred_margin
            home_win_prob = model.margin_to_home_win_prob(pred_margin, residual_std)

            # Dub Beta Model (XGBoost) -- read predict_week.py's latest
            # predictions CSV rather than refitting here. predict_week.py is
            # its own earlier pipeline step (see run_pipeline.py's STEPS) and
            # already paid for the XGBoost hyperparameter search once for
            # this exact upcoming week; refitting again here would pay for
            # the same search a second time on every pipeline run for no
            # benefit. If that CSV is missing, stale (a different week), or
            # predates the "XGBoost Line (Home)" column, beta_by_game just
            # stays empty and every game below shows no Dub Beta line --
            # same graceful-degradation as a missing market line elsewhere
            # in this file, never a crash.
            beta_by_game = {}
            pred_path = latest_predictions_file()
            if pred_path is not None:
                beta_csv = pd.read_csv(pred_path)
                if "XGBoost Line (Home)" in beta_csv.columns:
                    beta_by_game = dict(zip(
                        beta_csv["Game ID"].astype(int), beta_csv["XGBoost Line (Home)"]
                    ))

            # See src/totals_model.py -- SP+/PPA-based regression, same swap
            # made in backtest.py/predict_week.py, replacing the old
            # raw-scoring-average baseline (model.totals_baseline()).
            totals_train = totals_model.load_totals_training_frame(con)
            total_pipe, _ = totals_model.fit_total_model(totals_train)
            wk_totals_features = totals_model.load_upcoming_totals_frame(con, season, week)
            total_map = {}
            if not wk_totals_features.empty:
                total_preds = totals_model.predict_total(total_pipe, wk_totals_features)
                total_map = dict(zip(wk_totals_features["game_id"].astype(int), total_preds))

            m_by_id = {m["game_id"]: m for m in matchups}
            for i, row in enumerate(upcoming.itertuples()):
                mkt = m_by_id.get(row.game_id, {})
                market_spread_home = mkt.get("market_spread_home")
                market_total = mkt.get("market_total")
                edge = (market_spread_home - model_spread_home[i]) if market_spread_home is not None else None
                note_key = f"{row.away_team} @ {row.home_team}"
                model_total = total_map.get(int(row.game_id))

                # Once the game's final -- grade the model's own pick two
                # ways, same idea as backtest.py's bet grading but against
                # the model's own predicted margin rather than the market
                # line. straight_up: did the team the model favored just win
                # the game outright. covered_model_line: did the actual
                # margin fall on the model's side of its OWN predicted
                # margin (a stricter bar -- the favorite can win outright and
                # still "miss" its own number). Both are None until the game
                # is final, and null a "pick'em" (model_spread_home==0) up on
                # covered_model_line specifically since there's no favored
                # side to grade against.
                completed = bool(mkt.get("completed"))
                home_pts, away_pts = mkt.get("home_points"), mkt.get("away_points")
                straight_up_correct = covered_model_line = None
                actual_margin = None
                if completed and home_pts is not None and away_pts is not None:
                    actual_margin = home_pts - away_pts
                    pred_margin_i = float(pred_margin[i])
                    if pred_margin_i != 0:
                        straight_up_correct = (actual_margin > 0) == (pred_margin_i > 0)
                        diff = actual_margin - pred_margin_i
                        covered_model_line = (diff >= 0) if pred_margin_i > 0 else (diff <= 0)
                    elif actual_margin != 0:
                        straight_up_correct = None  # model called a dead-even pick'em; no favorite to grade

                # Dub Beta Model grading -- same two-way grade as the live
                # model just above (straight up / vs. its own line), plus
                # whether it agrees with the live model on which side it
                # favors. All three stay None whenever there's no Dub Beta
                # line for this game (see beta_by_game above) or, for the
                # agreement check, whenever either model calls a dead-even
                # pick'em (0.0) -- there's no "side" to compare in that case.
                beta_spread_home = beta_by_game.get(int(row.game_id))
                beta_straight_up_correct = beta_covered_model_line = None
                beta_agrees = None
                if beta_spread_home is not None and pd.notna(beta_spread_home):
                    beta_spread_home = float(beta_spread_home)
                    if completed and actual_margin is not None:
                        beta_margin_i = -beta_spread_home
                        if beta_margin_i != 0:
                            beta_straight_up_correct = (actual_margin > 0) == (beta_margin_i > 0)
                            beta_diff = actual_margin - beta_margin_i
                            beta_covered_model_line = (beta_diff >= 0) if beta_margin_i > 0 else (beta_diff <= 0)
                    if model_spread_home[i] != 0 and beta_spread_home != 0:
                        # bool(...) wrapper is load-bearing, not defensive
                        # style: model_spread_home[i] is a numpy scalar (the
                        # model's array output), so this comparison would
                        # otherwise produce numpy.bool_ -- which json.dumps()
                        # can't serialize (and which numpy 2.x's renamed
                        # scalar type displays as plain "bool" in the
                        # resulting TypeError, easy to misread as a real bug
                        # in a plain Python bool). Every other new field here
                        # already routes through a float()/bool() cast for
                        # exactly this reason; this was the one that didn't.
                        beta_agrees = bool((model_spread_home[i] < 0) == (beta_spread_home < 0))
                else:
                    beta_spread_home = None

                predictions.append({
                    "game_id": int(row.game_id), "week": int(row.week),
                    "away_team": row.away_team, "home_team": row.home_team,
                    "model_spread_home": round(model_spread_home[i], 1),
                    "model_total": round(model_total, 1) if model_total is not None else None,
                    # Market/Vegas line, shown directly on the Predictions
                    # page next to the model's own line -- previously only
                    # the derived "edge" (market minus model) was exposed,
                    # never the raw market number itself.
                    "market_spread_home": round(market_spread_home, 1) if market_spread_home is not None else None,
                    "market_total": round(market_total, 1) if market_total is not None else None,
                    "home_win_prob": round(float(home_win_prob[i]), 3),
                    "edge": round(edge, 1) if edge is not None else None,
                    "notes": notes.get(note_key, ""),
                    "completed": completed,
                    "home_points": home_pts,
                    "away_points": away_pts,
                    "straight_up_correct": straight_up_correct,
                    "covered_model_line": covered_model_line,
                    # Dub Beta Model (XGBoost) -- informational only, see
                    # this function's beta_by_game comment above. Never used
                    # to compute edge/is_bet -- only the live model's "edge"
                    # field above ever decides that.
                    "beta_model_spread_home": round(beta_spread_home, 1) if beta_spread_home is not None else None,
                    "beta_agrees": beta_agrees,
                    "beta_straight_up_correct": beta_straight_up_correct,
                    "beta_covered_model_line": beta_covered_model_line,
                })

    return matchups, predictions, {"season": season, "week": week}


def _bet_type_block(summary: dict, win_rate_key: str) -> dict:
    n_bets = summary.get("n_bets", 0) or 0
    return {
        "n_bets": n_bets,
        "wins": summary.get("wins", 0),
        "losses": summary.get("losses", 0),
        "pushes": summary.get("pushes", 0),
        "win_rate": summary.get(win_rate_key),
        "units": summary.get("units_won") if "units_won" in summary else round((summary.get("roi_flat_stake") or 0) * n_bets, 2),
    }


def build_model_tracking(con):
    df = backtest.run_backtest(con)
    if df.empty:
        return {"n_games": 0, "note": "Not enough graded history yet."}

    season_df = df[df["season"] == CURRENT_SEASON]
    mw_df = season_df[season_df["is_mw_game"]]

    spread = backtest.summarize(mw_df, "spread")
    total = backtest.summarize_totals(mw_df, "total")
    moneyline = backtest.summarize_moneyline(mw_df, "moneyline")

    bets = mw_df[mw_df["is_bet"] & mw_df["bet_result"].isin(["Win", "Loss"])]
    upsets = bets[bets.apply(
        lambda r: (r["lean"] == "Home" and r["market_spread_home"] > 0)
        or (r["lean"] == "Away" and r["market_spread_home"] < 0), axis=1
    )] if not bets.empty else bets
    upset_wins = int((upsets["bet_result"] == "Win").sum()) if not upsets.empty else 0
    upset_total = len(upsets)

    return {
        "season": CURRENT_SEASON,
        "n_games": spread.get("n_games", 0),
        "spread": _bet_type_block(spread, "ats_win_rate"),
        "total": _bet_type_block(total, "win_rate"),
        "moneyline": _bet_type_block(moneyline, "win_rate"),
        "upset_calls": upset_total,
        "upset_wins": upset_wins,
        "note": "Hypothetical flat 1-unit-per-bet tracking of the model's own picks (spread/total at -110, "
                "moneyline at the real posted odds) -- not your personal bets.",
    }


def build_beta_tracking():
    """
    Dub Beta Model (XGBoost)'s season-to-date -- spread only (see this
    module's docstring for why: model_comparison.py, the source of this
    data, only ever compares margin models). Sourced from
    data/clean/model_comparison_results.csv via load_beta_comparison(), so
    this is a snapshot from the last time you ran `python
    src/model_comparison.py` by hand, not something recomputed on every
    pipeline run -- same staleness contract as the Results page's beta_model
    field and excel/update_model_comparison_tab.py's Upcoming Games section.
    """
    df = load_beta_comparison()
    if df is None:
        return {"n_games": 0, "note": "No Dub Beta Model comparison yet -- run `python src/model_comparison.py`, "
                                       "then `python src/export_site_data.py`, to populate this."}

    season_df = df[(df["season"] == CURRENT_SEASON) & df["is_mw_game"]]
    if season_df.empty:
        return {"n_games": 0, "note": f"No {CURRENT_SEASON} Mountain West games graded in the Dub Beta "
                                       "comparison yet."}

    bets = season_df[season_df["xgb_is_bet"] & season_df["xgb_result"].isin(["Win", "Loss"])]
    wins = int((bets["xgb_result"] == "Win").sum())
    losses = int((bets["xgb_result"] == "Loss").sum())
    pushes = int((season_df["xgb_is_bet"] & (season_df["xgb_result"] == "Push")).sum())
    n_bets = wins + losses
    profit = wins * (100 / 110) - losses * 1.0

    agree = season_df["models_agree"].dropna()
    agree_rate = float(agree.mean()) if not agree.empty else None

    return {
        "season": CURRENT_SEASON,
        "n_games": len(season_df),
        "spread": {
            "n_bets": n_bets, "wins": wins, "losses": losses, "pushes": pushes,
            "win_rate": round(wins / n_bets, 4) if n_bets else None,
            "units": round(profit, 2) if n_bets else 0.0,
        },
        "agree_rate": round(agree_rate, 4) if agree_rate is not None else None,
        # No "note" here on purpose (unlike the two empty-data returns above,
        # where it's actionable guidance) -- the section header's "Candidate,
        # Spread Only" badge already says what this card is; a restated
        # paragraph under real numbers was just clutter. See tracking.html's
        # renderBeta(), which only renders a notes-box when .note is present.
        "note": None,
    }


def _beta_result_dict(b):
    """b is one row (itertuples) from model_comparison_results.csv, or None
    if this game isn't in that snapshot. See build_results()'s docstring."""
    if b is None:
        return None
    return {
        "spread_pick": round(float(b.xgb_spread_home), 1) if pd.notna(b.xgb_spread_home) else None,
        "spread_lean": b.xgb_lean if isinstance(b.xgb_lean, str) else None,
        "spread_was_bet": bool(b.xgb_is_bet),
        # xgb_result round-trips through the CSV as a real string or NaN
        # (never bet -> Python None -> pandas NaN, same trap this module's
        # _sanitize_nans() exists for elsewhere) -- guard explicitly rather
        # than relying on isinstance() alone catching it.
        "spread_result": b.xgb_result if isinstance(b.xgb_result, str) and pd.notna(b.xgb_result) else None,
        "agrees_with_model": bool(b.models_agree) if pd.notna(b.models_agree) else None,
    }


def build_matchup_grid(con):
    """
    Matchup Creator page: precomputes the live Ridge model's predicted
    spread/win-probability for every ORDERED pair of the 10 current (2026)
    Mountain West teams (home, away), including pairs that aren't actually
    on this season's schedule -- e.g. "what would Boise State -7 at home vs.
    Wyoming look like" even if they aren't playing each other this year. The
    site is fully static (GitHub Pages, no server/API), so this has to be
    precomputed here, not looked up client-side -- same "everything
    precomputed, JS just renders" pattern as every other page.

    Refits the SAME Ridge pipeline (model.py's build_pipeline()/
    fit_margin_model(), on ALL available completed games, no walk-forward
    cutoff) that predict_week.py fits for the live Weekly Slate -- cheap
    (RidgeCV, no hyperparameter search), so refitting it here instead of
    trying to reuse predict_week.py's in-memory pipe (which isn't persisted
    anywhere) costs one extra Ridge fit per pipeline run, not a real backtest
    or a search. This intentionally does NOT touch XGBoost/Dub Beta at all --
    that model has no matchup-predictor role anywhere on the site.

    FEATURE_COLS is mostly team-season-level lookups that are perfectly safe
    to reuse for a hypothetical, unscheduled matchup: rating_diff (today's
    Elo via power_rating.current_ratings(), already computed above for
    build_rankings()), sp_diff/ppa_diff (prior-season SP+/net-PPA),
    talent_diff (this season's recruiting), and the 7 drive_*_diff fields
    (prior-season rates, via drive_stats_snapshots' last as_of_week). Only
    five fields are genuinely game-specific/situational and need an assumed
    value for a hypothetical, since there's no real scheduled game to read
    them from:
      - rest_diff = 0 (assume both teams equally rested)
      - neutral_site_flag = 0 (assume it's played at the "home" team's own
        usual venue -- see travel/elevation below)
      - conference_game_flag = 1 and mw_involved_flag = 1 (both true for
        any pair drawn from MW_TEAMS_2026 only, which is this grid's whole
        scope)
      - travel_km_away / elevation_delta_away_ft: NOT assumed -- these ARE
        still computable for a hypothetical, using each team's real usual
        home venue (features.py's own team_home_venues()/haversine_km()):
        the "home" team's usual venue is assumed to host it, and the "away"
        team's travel/elevation change is measured from ITS OWN usual venue
        to that game site, exactly like a real scheduled game would be.

    Any lookup a team is missing (most commonly drive_stats_snapshots, which
    can be entirely empty on an older database -- see features.py's own
    defensive try/except) is left as None/NaN in the feature row rather than
    guessed at; the pipeline's own SimpleImputer(strategy="median"), fit on
    the real training data, fills it in exactly like it would for a real
    game missing that same feature, so a team with no drive-stats history
    still gets a reasonable (if less-informed) prediction instead of a
    crash.

    Returns {"teams": [...], "grid": {home: {away: {...}}}} -- see the
    per-pair dict below for the exact fields.
    """
    train_df = model.load_training_frame(con)
    pipe, residual_std = model.fit_margin_model(train_df)

    teams = sorted(MW_TEAMS_2026)
    prior_season = CURRENT_SEASON - 1

    ratings_now = current_ratings(con)
    sp_map = {(r.season, r.team): r.rating for r in con.execute(
        "SELECT season, team, rating FROM sp_ratings"
    ).fetchdf().itertuples()}
    ppa_map = {
        (r.season, r.team): r.off_ppa - r.def_ppa
        for r in con.execute(
            "SELECT season, team, off_ppa, def_ppa FROM advanced_stats"
        ).fetchdf().itertuples()
        if pd.notna(r.off_ppa) and pd.notna(r.def_ppa)
    }
    talent_map = {(r.season, r.team): r.points for r in con.execute(
        "SELECT season, team, points FROM recruiting"
    ).fetchdf().itertuples()}

    try:
        drive_stats_df = con.execute("""
            SELECT ds.season, ds.team, ds.yards_per_drive, ds.points_per_drive,
                   ds.turnovers_per_drive, ds.pass_yards_per_drive, ds.rush_yards_per_drive,
                   ds.yards_per_attempt, ds.yards_per_carry
            FROM drive_stats_snapshots ds
            INNER JOIN (
                SELECT season, team, MAX(as_of_week) AS max_week
                FROM drive_stats_snapshots GROUP BY season, team
            ) mx ON mx.season = ds.season AND mx.team = ds.team AND mx.max_week = ds.as_of_week
        """).fetchdf()
        drive_map = {(r.season, r.team): r for r in drive_stats_df.itertuples()}
    except duckdb.Error:
        drive_map = {}

    games = con.execute("""
        SELECT home_team, away_team, venue_id, neutral_site FROM games
    """).fetchdf()
    home_venue = team_home_venues(games) if not games.empty else {}
    venues = con.execute("SELECT venue_id, latitude, longitude, elevation_ft FROM venues").fetchdf()
    venues = venues.set_index("venue_id") if not venues.empty else venues

    def venue_for(team):
        vid = home_venue.get(team)
        if vid is None or venues is None or venues.empty or vid not in venues.index:
            return None, None, None
        row = venues.loc[vid]
        return row["latitude"], row["longitude"], row["elevation_ft"]

    def _drive_diff(home_drive, away_drive, field):
        if home_drive is None or away_drive is None:
            return None
        h, a = getattr(home_drive, field), getattr(away_drive, field)
        if h is None or a is None or pd.isna(h) or pd.isna(a):
            return None
        return h - a

    rows, pair_keys = [], []
    for home in teams:
        home_rating = ratings_now.get(home)
        home_sp, home_ppa = sp_map.get((prior_season, home)), ppa_map.get((prior_season, home))
        home_talent = talent_map.get((CURRENT_SEASON, home))
        home_drive = drive_map.get((prior_season, home))
        home_lat, home_lon, home_elev = venue_for(home)

        for away in teams:
            if away == home:
                continue
            away_rating = ratings_now.get(away)
            away_sp, away_ppa = sp_map.get((prior_season, away)), ppa_map.get((prior_season, away))
            away_talent = talent_map.get((CURRENT_SEASON, away))
            away_drive = drive_map.get((prior_season, away))
            away_lat, away_lon, away_elev = venue_for(away)

            rating_diff = (home_rating - away_rating) if (home_rating is not None and away_rating is not None) else None
            sp_diff = (home_sp - away_sp) if (home_sp is not None and away_sp is not None) else None
            ppa_diff = (home_ppa - away_ppa) if (home_ppa is not None and away_ppa is not None) else None
            talent_diff = (home_talent - away_talent) if (home_talent is not None and away_talent is not None) else None
            travel_km_away = haversine_km(home_lat, home_lon, away_lat, away_lon)
            elevation_delta_away = (home_elev - away_elev) if (home_elev is not None and away_elev is not None) else None

            rows.append({
                "rating_diff": rating_diff, "rest_diff": 0.0,
                "travel_km_away": travel_km_away,
                "elevation_delta_away_ft": elevation_delta_away,
                "neutral_site_flag": 0.0, "conference_game_flag": 1.0,
                "sp_diff": sp_diff, "ppa_diff": ppa_diff, "talent_diff": talent_diff,
                "mw_involved_flag": 1.0,
                "drive_yards_diff": _drive_diff(home_drive, away_drive, "yards_per_drive"),
                "drive_points_diff": _drive_diff(home_drive, away_drive, "points_per_drive"),
                "drive_turnovers_diff": _drive_diff(home_drive, away_drive, "turnovers_per_drive"),
                "pass_ypd_diff": _drive_diff(home_drive, away_drive, "pass_yards_per_drive"),
                "rush_ypd_diff": _drive_diff(home_drive, away_drive, "rush_yards_per_drive"),
                "ypa_diff": _drive_diff(home_drive, away_drive, "yards_per_attempt"),
                "ypc_diff": _drive_diff(home_drive, away_drive, "yards_per_carry"),
            })
            pair_keys.append((home, away))

    grid = {home: {} for home in teams}
    if rows:
        feat_df = pd.DataFrame(rows, columns=model.FEATURE_COLS)
        pred_margin = model.predict_margin(pipe, feat_df)
        spread_home = model.margin_to_model_spread_home(pred_margin)
        win_prob = model.margin_to_home_win_prob(pred_margin, residual_std)
        for (home, away), margin, spread, prob in zip(pair_keys, pred_margin, spread_home, win_prob):
            grid[home][away] = {
                "predicted_margin": round(float(margin), 1),
                "spread_home": round(float(spread), 1),
                "home_win_prob": round(float(prob), 4),
            }

    return {"teams": teams, "grid": grid, "residual_std": round(residual_std, 2)}


def build_results(con):
    """
    Past Results page: every completed MW game this season, the model's own
    graded pick (spread/total/moneyline -- same walk-forward grading as
    backtest.py/the Tracking page, not just the ones that cleared the bet
    threshold) side by side with your own placed bets on that same game, if
    any. Matched to your Bet Log by the exact "Away @ Home" text in column C
    -- see read_bet_log()'s docstring.

    Also carries a "beta_model" field per game -- the Dub Beta Model
    (XGBoost)'s own graded spread pick for that same game, from
    load_beta_comparison(). null whenever that game isn't in the last
    model_comparison.py snapshot (script never run, or run before this game
    was graded) -- see this module's docstring for the staleness contract.
    """
    df = backtest.run_backtest(con)
    if df.empty:
        return []
    season_df = df[(df["season"] == CURRENT_SEASON) & df["is_mw_game"]]
    if season_df.empty:
        return []

    game_ids = [int(g) for g in season_df["game_id"].tolist()]
    placeholders = ",".join("?" * len(game_ids))
    meta = con.execute(f"""
        SELECT game_id, start_date, home_points, away_points
        FROM games WHERE game_id IN ({placeholders})
    """, game_ids).fetchdf().set_index("game_id")

    bets_by_matchup = {}
    for b in read_bet_log().get("graded", []):
        bets_by_matchup.setdefault(b["matchup"], []).append(b)

    beta_df = load_beta_comparison()
    beta_by_game = {int(r.game_id): r for r in beta_df.itertuples()} if beta_df is not None else {}

    def _n(v):
        return round(float(v), 1) if pd.notna(v) else None

    results = []
    for row in season_df.itertuples():
        gid = int(row.game_id)
        m = meta.loc[gid] if gid in meta.index else None
        matchup_key = f"{row.away_team} @ {row.home_team}"
        results.append({
            "game_id": gid, "season": int(row.season), "week": int(row.week),
            "date": pd.to_datetime(m["start_date"]).strftime("%Y-%m-%d") if m is not None else None,
            "away_team": row.away_team, "home_team": row.home_team,
            "home_points": int(m["home_points"]) if m is not None and pd.notna(m["home_points"]) else None,
            "away_points": int(m["away_points"]) if m is not None and pd.notna(m["away_points"]) else None,
            "model": {
                "spread_pick": _n(row.model_spread_home),
                "market_spread_close": _n(row.market_spread_home),
                "spread_lean": row.lean, "spread_was_bet": bool(row.is_bet),
                "spread_result": row.bet_result,
                "total_pick": _n(row.model_total),
                "market_total_close": _n(row.market_total),
                "total_lean": row.total_lean,
                "total_was_bet": bool(row.is_total_bet) if row.is_total_bet is not None else None,
                "total_result": row.total_bet_result,
                "ml_lean": row.ml_lean,
                "ml_was_bet": bool(row.is_ml_bet) if row.is_ml_bet is not None else None,
                "ml_result": row.ml_bet_result,
            },
            "beta_model": _beta_result_dict(beta_by_game.get(gid)),
            "your_bets": bets_by_matchup.get(matchup_key, []),
        })

    results.sort(key=lambda r: (r["date"] or "", r["week"]), reverse=True)
    return results


def _cell_date_str(value):
    """Bet Log's Date column may hold a real date/datetime (typed into Excel
    as a date) or plain text (e.g. "2026-08-29") -- normalize either to a
    plain ISO-ish string for display, without needing pandas involved."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def read_bet_log(tracker_path: Path = TRACKER_PATH):
    """
    Your actual placed bets, straight from the Bet Log tab -- read fresh from
    the raw input cells (Odds, Stake, Result) and re-computed with the same
    odds math as the model's own grading, rather than trusting the sheet's
    cached formula values (openpyxl never recalculates formulas itself, so a
    cached value is only as fresh as the last time the file was opened in
    real Excel).

    Also surfaces PENDING bets -- rows where you've filled in the bet itself
    (Matchup/Bet Type/Side/Odds/Stake) but haven't typed a Result yet, i.e.
    action you've already placed that just hasn't been graded. Those are
    listed separately and never counted in the win/loss record below.

    "graded" carries every individual graded row (not just the by_type
    aggregates above) -- the Results page (build_results()) matches these
    back to a specific game by the exact "Away @ Home" matchup text you
    typed into column C, the same convention site_notes.json/
    manual_lines.json already use elsewhere in this file.
    """
    empty = {"n_bets": 0, "pending": [], "graded": []}
    if not tracker_path.exists():
        return {**empty, "note": "Tracker workbook not found -- run excel/build_tracker.py first."}

    wb = openpyxl.load_workbook(tracker_path, data_only=False)
    if "Bet Log" not in wb.sheetnames:
        return {**empty, "note": "No Bet Log tab found in the tracker workbook."}
    ws = wb["Bet Log"]

    by_type = {}    # bet type -> {wins, losses, pushes, units}
    pending = []    # bets placed but not yet graded (no Result typed in)
    graded = []     # every individual graded row, for per-game matching
    for r in BET_LOG_ROWS:
        bet_type = ws[f"D{r}"].value
        odds = ws[f"G{r}"].value
        stake = ws[f"H{r}"].value
        result = ws[f"K{r}"].value
        if not bet_type or odds is None or stake is None:
            continue

        if not result:
            pending.append({
                "date": _cell_date_str(ws[f"A{r}"].value),
                "week": ws[f"B{r}"].value,
                "matchup": ws[f"C{r}"].value,
                "bet_type": str(bet_type).strip(),
                "side": ws[f"E{r}"].value,
                "odds": float(odds),
                "stake": float(stake),
            })
            continue

        result = str(result).strip().upper()
        if result not in ("W", "L", "P"):
            continue

        bucket = by_type.setdefault(str(bet_type).strip(), {"wins": 0, "losses": 0, "pushes": 0, "units": 0.0})
        won = result == "W"
        units = 0.0
        if result == "P":
            bucket["pushes"] += 1
        else:
            bucket["wins" if won else "losses"] += 1
            units = payout_profit(float(stake), float(odds), won)
            bucket["units"] += units

        graded.append({
            "date": _cell_date_str(ws[f"A{r}"].value),
            "week": ws[f"B{r}"].value,
            "matchup": ws[f"C{r}"].value,
            "bet_type": str(bet_type).strip(),
            "side": ws[f"E{r}"].value,
            "odds": float(odds), "stake": float(stake),
            "result": result, "units": round(units, 2),
        })

    pending.sort(key=lambda b: b["date"] or "")

    if not by_type:
        return {**empty, "pending": pending,
                "note": "No graded bets in the Bet Log yet -- add a Result (W/L/P) to see your record here."}

    by_type_out = {}
    total_wins = total_losses = total_pushes = 0
    total_units = 0.0
    for bet_type, b in by_type.items():
        n_bets = b["wins"] + b["losses"]
        by_type_out[bet_type] = {
            "wins": b["wins"], "losses": b["losses"], "pushes": b["pushes"],
            "win_rate": round(b["wins"] / n_bets, 4) if n_bets else None,
            "units": round(b["units"], 2),
        }
        total_wins += b["wins"]
        total_losses += b["losses"]
        total_pushes += b["pushes"]
        total_units += b["units"]

    total_n_bets = total_wins + total_losses
    return {
        "n_bets": total_n_bets,
        "wins": total_wins, "losses": total_losses, "pushes": total_pushes,
        "win_rate": round(total_wins / total_n_bets, 4) if total_n_bets else None,
        "units": round(total_units, 2),
        "by_type": by_type_out,
        "pending": pending,
        "graded": graded,
        "note": "Your actual bets from the Bet Log tab, graded at the real odds you entered.",
    }


def main():
    DOCS_DATA.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(str(DB_PATH))
    notes = load_notes()
    manual_lines = load_manual_lines()

    rankings = build_rankings(con)
    matchups, predictions, week_info = build_matchups_and_predictions(con, notes, manual_lines)
    live_lines, lines_week_info = build_live_lines(con, manual_lines)
    line_history = build_line_history(con, [g["game_id"] for g in live_lines])
    tracking = {"model": build_model_tracking(con), "yours": read_bet_log(), "beta_model": build_beta_tracking()}
    results = build_results(con)
    matchup_grid = build_matchup_grid(con)
    con.close()

    meta = {"generated_at": _now_iso(), "current_week": week_info}
    lines_meta = {"generated_at": _now_iso(), "current_week": lines_week_info}

    def _write(name, payload):
        (DOCS_DATA / name).write_text(json.dumps(_sanitize_nans(payload), indent=2))

    _write("rankings.json", {"meta": meta, "rankings": rankings})
    _write("matchups.json", {"meta": meta, "matchups": matchups})
    _write("predictions.json", {"meta": meta, "predictions": predictions})
    _write("tracking.json", {"meta": meta, "tracking": tracking})
    _write("lines.json", {"meta": lines_meta, "games": live_lines})
    _write("line_history.json", {"meta": lines_meta, "history": line_history})
    _write("results.json", {"meta": meta, "results": results})
    _write("matchup_grid.json", {"meta": meta, "teams": matchup_grid["teams"],
                                  "grid": matchup_grid["grid"], "residual_std": matchup_grid["residual_std"]})

    print(f"Wrote rankings ({len(rankings)}), matchups ({len(matchups)}), "
          f"predictions ({len(predictions)}), live lines ({len(live_lines)}), "
          f"results ({len(results)}), tracking summary, matchup grid "
          f"({len(matchup_grid['teams'])} teams) to {DOCS_DATA}")


if __name__ == "__main__":
    main()
